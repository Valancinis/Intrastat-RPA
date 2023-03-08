from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.edge.service import Service
import openpyxl
from openpyxl import load_workbook
from selenium.webdriver import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException

# Set up the driver
s = Service('C:/Users/dovydas.valancius/Edge Web Drivers/edgedriver_win64/msedgedriver.exe')
driver = webdriver.Edge(service=s)

# Load the workbook and select the worksheet
workbook = load_workbook(filename="intrastat for RPA.xlsx")
worksheet = workbook.active

# Create action chain object
action = ActionChains(driver)

# Take login info from a txt file
with open("logins.txt", "r") as file:
    # read the lines of the file
    lines = file.readlines()

# extract the username and password from the file contents
user_name = lines[0].strip().split(':')[-1]
pass_word = lines[1].strip().split(':')[-1]

# Navigate to the login page
driver.get("http://lsge1pd.amer.thermo.com/jde/E1Menu.maf?jdeLoginAction=LOGOUT&RENDER_MAFLET=E1Menu")

# Find the user ID input field and enter the value
user_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "User")))
user_input.click()
user_input.send_keys(user_name)

# Find the password input field and enter the value
password_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "Password")))
password_input.click()
password_input.send_keys(pass_word)

# Find the "Sign In" input field and click it
signin_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@value='Sign In']")))
signin_button.click()

# Find the "OK" input field and click it in Select Env. Page
ok_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@id='SUBMIT_BUTTON']")))
ok_button.click()

# Wait for the intrastat window element to be available
intrastat_window = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,
                                                                                   "//div[@class='listItem' and @id='listFav_17099562200007_IUGMOHNGYB_HLXSFWFDFS_07____500_13354848600006_IUGMOHNGYB_PPFINWYYZY_APP_P0018T_W0018TA_null_70']")))
time.sleep(1)

# Click the intrastat window element
intrastat_window.click()

# Switch to the iframe
iframe = WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "e1menuAppIframe")))

# Map all the needed columns
column_mapping = {
    'Order Number': 1,
    'Line Number': 2,
    'Order type' : 3,
    'Country Of Original Origin': 4,
    'Commodity Code': 5,
    'Net Mass in KG': 6,
    'Mode of Transportation': 7,
    'Conditions of Transportation': 8,
    'Supplementary Units': 9,
}


# Function that fills in E1 fields
def update_information():
    # Enter country of origin
    country_input = worksheet.cell(row=update_row + 2, column=column_mapping['Country Of Original Origin']).value
    if country_input and str(country_input).strip():
        country_input = str(country_input).replace(" ", "")
        codes_tab = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//span[@id='outerJDETabHeader0_28']//td//td[3]//td[2]//table")))
        codes_tab.click()
        enter_value('C0_189', country_input)

    # Enter mode of transportation
    mode_input = worksheet.cell(row=update_row + 2, column=column_mapping['Mode of Transportation']).value
    if mode_input and str(mode_input).strip():
        mode_input = str(mode_input).replace(" ", "")
        codes_tab = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//span[@id='outerJDETabHeader0_28']//td//td[3]//td[2]//table")))
        codes_tab.click()
        enter_value('C0_193', mode_input)

    # Enter conditions of transportation
    condition_input = worksheet.cell(row=update_row + 2, column=column_mapping['Conditions of Transportation']).value
    if condition_input and str(condition_input).strip():
        condition_input = str(condition_input).replace(" ", "")
        codes_tab = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//span[@id='outerJDETabHeader0_28']//td//td[3]//td[2]//table")))
        codes_tab.click()
        enter_value('C0_195', condition_input)

    # Enter commodity codes
    commodity_input = worksheet.cell(row=update_row + 2, column=column_mapping['Commodity Code']).value
    if commodity_input and str(commodity_input).strip():
        commodity_input = str(commodity_input).replace(" ", "")
        codes_tab = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//span[@id='outerJDETabHeader0_28']//td//td[3]//td[2]//table")))
        codes_tab.click()
        enter_value('C0_258', commodity_input)

    # Enter Net Mass in KG values
    weight_input = worksheet.cell(row=update_row + 2, column=column_mapping['Net Mass in KG']).value
    if weight_input and str(weight_input).strip():
        weight_input = str(weight_input).replace(" ", "")
        amounts_tab = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//span[@id='outerJDETabHeader0_28']//td//td[2]//td[2]//table")))
        amounts_tab.click()
        enter_value('C0_262', weight_input)

    # Enter Supplementary Unit values
    unit_input = worksheet.cell(row=update_row + 2, column=column_mapping['Supplementary Units']).value
    if unit_input and str(unit_input).strip():
        unit_input = str(unit_input).replace(" ", "")
        amounts_tab = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//span[@id='outerJDETabHeader0_28']//td//td[2]//td[2]//table")))
        amounts_tab.click()
        enter_value('C0_276', unit_input)


# Custom function to update each value box in E1
def enter_value(field_id, value):
    field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, field_id)))
    action.click(field)
    action.key_down(Keys.CONTROL)
    action.send_keys("a")
    action.key_up(Keys.CONTROL)
    action.send_keys(value)
    action.perform()


# Which Excel file row is being updated?
current_row = 1
print(f"Started updating line:{current_row}")

# Define a variable that will be incremented and used to pick a row in Excel file to pull data from
update_row = 0

# Begin a loop that goes through each excel line and updates information
for row in range(2, worksheet.max_row + 1):

    # Checks if there are more lines in Excel file
    if worksheet.cell(row=update_row + 2, column=1).value is not None:

        # Wait for the page to load
        time.sleep(0.5)

        # Enter order number
        order_number = worksheet.cell(row=update_row + 2, column=column_mapping['Order Number']).value
        enter_value('C0_19', order_number)

        # Enter line number
        line_value = worksheet.cell(row=update_row + 2, column=column_mapping['Line Number']).value
        enter_value('C0_25', line_value)

        # Enter order type
        line_value = worksheet.cell(row=update_row + 2, column=column_mapping['Order type']).value
        field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@name='qbe0_1.2']")))
        action.click(field)
        action.key_down(Keys.CONTROL)
        action.send_keys("a")
        action.key_up(Keys.CONTROL)
        action.send_keys(line_value)
        action.perform()

        # Click search button
        search_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//a[@id='C0_11']")))
        search_button.click()

        # Wait for the page to load
        time.sleep(0.5)

        # Get a number of rows in the search results in E1
        rows = WebDriverWait(driver, 10).until(
            EC.visibility_of_all_elements_located((By.XPATH, "//table[@id='jdeGridData0_1']//input")))

        # Begin another loop that goes through each search result in E1 and updates it
        for index, line in enumerate(rows):
            checkboxes = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,
                                                                                           "//table[@id='jdeGridData0_1']//tr[@id='G0_1_R" + str(
                                                                                               index) + "']//input")))  # ČIA PROBLEMA YRA SU str(line)
            action.double_click(checkboxes).perform()

            # Run the function to update values in a record
            update_information()

            # Wait for error to appear and check if input raise any errors in the system
            time.sleep(0.5)
            try:
                # Wait for error message to appear
                error_msg_element = WebDriverWait(driver, 1).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "InYourFaceError")))

                # Print error message and row number
                print("Code has stopped because there is an unknown value in excel row:", current_row)

                # Exit loop
                break

            # If timeout exception happens
            except TimeoutException:

                # Error message did not appear, continue with next row
                current_row += 1

                # Click OK button
                ok_button = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@id='C0_11']")))
                ok_button.click()

                # Find the "Save" button and press it
                save_button = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@id='C0_11']")))
                save_button.click()

                # Wait for the page to load.
                time.sleep(0.5)

                # Find the Exit button, press it, print line number that was just updated and increment the update_row variable
                exit_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@id='C0_12']")))
                exit_button.click()
                time.sleep(1)
                print(f"Currently updating row:{current_row}")
                update_row += 1
                continue

print("Completed")
# Wait for 2 seconds before quitting the program
time.sleep(2)
# Close the browser
driver.quit()
