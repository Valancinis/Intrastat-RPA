from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.edge.service import Service
import openpyxl
from openpyxl import load_workbook
from selenium.webdriver import ActionChains

# Set up the driver
s = Service('C:/Users/dovydas.valancius/Edge Web Drivers/edgedriver_win64/msedgedriver.exe')
driver = webdriver.Edge(service=s)

# Load the workbook and select the worksheet
workbook = load_workbook(filename="intrastat for RPA.xlsx")
worksheet = workbook.active

# Create action chain object
action = ActionChains(driver)

# Take login info from a txt file
with open("logins.txt", "r") as file:

    # read the lines of the file
    lines = file.readlines()

# extract the username and password from the file contents
user_name = lines[0].strip().split(':')[-1]
pass_word = lines[1].strip().split(':')[-1]

# Navigate to the login page
driver.get("http://lsge1pd.amer.thermo.com/jde/E1Menu.maf?jdeLoginAction=LOGOUT&RENDER_MAFLET=E1Menu")

# Find the user ID input field and enter the value
user_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "User")))
user_input.click()
user_input.send_keys(user_name)

# Find the password input field and enter the value
password_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "Password")))
password_input.click()
password_input.send_keys(pass_word)

# Find the "Sign In" input field and click it
signin_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@value='Sign In']")))
signin_button.click()

# Find the "OK" input field and click it in Select Env. Page
ok_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@id='SUBMIT_BUTTON']")))
ok_button.click()

# Wait for the intrastat window element to be available
intrastat_window = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@class='listItem' and @id='listFav_17099562200007_IUGMOHNGYB_HLXSFWFDFS_07____500_60DDB18A000601_00000CA400000B08_FAJW_APP_P584101_W584101A_null_170']")))
time.sleep(1)
# Click the intrastat window element
intrastat_window.click()

# Switch to the iframe
iframe = WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "e1menuAppIframe")))

# Which Excel file row is being updated?
current_row = 1
print(f"Started updating line:{current_row}")

# loop through each row in the worksheet starting from row 2
for row in range(2, worksheet.max_row + 1):

    # Checks if there are more lines in Excel file
    if worksheet.cell(row=row, column=10).value is not None:

        # Find the "SKU number" input field and enter the value from the Excel file
        SKU_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "C0_34")))
        SKU_number = worksheet.cell(row=row, column=10).value
        action.click(SKU_input)
        action.key_down(Keys.CONTROL)
        action.send_keys("a")
        action.key_up(Keys.CONTROL)
        action.send_keys(SKU_number)
        action.perform()

        time.sleep(0.25)

        # Find the "Search Button" and press it
        search_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//a[@id='C0_15']")))
        search_button.click()

        time.sleep(0.25)

        # Find the "Check box" and double click it
        search_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//td[@class='JSSelectGrid  selectedModifier']//input")))
        action.double_click(search_button).perform()

        # Waiting for the page to load
        time.sleep(1)

        # Check if the "Country Of Original Origin" column has a value and write it in to E1
        country_input = worksheet.cell(row=row, column=4).value

        if country_input and str(country_input).strip():
            country_input = str(country_input).replace(" ", "")

            # Find the "Country Of Origin" input field and enter the value from the Excel file
            country_input_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//td[@colindex='4']//input")))
            action.click(country_input_field)
            action.key_down(Keys.CONTROL)
            action.send_keys("a")
            action.key_up(Keys.CONTROL)
            action.send_keys(country_input)
            action.perform()

        # Check if the "Commodity code" column has a value and writes it in to E1
        commodity_input = worksheet.cell(row=row, column=5).value

        if commodity_input and str(commodity_input).strip():
            commodity_input = str(commodity_input).replace(" ", "")

            # Find the "Commodity codes" input field and enter the value from the Excel file
            commodity_input_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//td[@colindex='5']//input")))
            action.click(commodity_input_field)
            action.key_down(Keys.CONTROL)
            action.send_keys("a")
            action.key_up(Keys.CONTROL)
            action.send_keys(commodity_input)
            action.perform()

        # Check if the "Net Weight" column has a value and writes it in to E1
        weight_input = worksheet.cell(row=row, column=6).value

        if weight_input and str(weight_input).strip():
            weight_input = str(weight_input).replace(" ", "")

            # Find the "Net Mass in KG" input field and enter the value from the Excel file
            weight_input_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//td[@colindex='6']//input")))
            action.click(weight_input_field)
            action.key_down(Keys.CONTROL)
            action.send_keys("a")
            action.key_up(Keys.CONTROL)
            action.send_keys(weight_input)
            action.perform()
        current_row += 1

        # Find the "Save" button and press it
        save_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//a[@id='C0_12']")))
        save_button.click()

        time.sleep(1)
        print(f"Finished updating row:{current_row}")

time.sleep(2)
print("Done")

# Close the browser
driver.quit()
